import os
import json
import requests
from datetime import datetime, timedelta, timezone
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import re as _re

# ── Configuration ──────────────────────────────────────────────────────────────
RECEIVER_EMAILS = os.environ.get('RECIPIENT_EMAIL')

CONFIG = {
    'GPS_API_KEY': os.environ.get('API_KEY'),
    'VEHICLES': {
        '868373075408486': '3922УБЯ',
        '350544501468303': '5034УКН',
        '863719068034074': '5035УКН',
        '350317174707566': '5036УКН'

    },
    'SENDER_EMAIL':   os.environ.get('GMAIL_EMAIL'),
    'SENDER_PASSWORD': os.environ.get('GMAIL_PASSWORD'),
    'RECEIVER_EMAILS': RECEIVER_EMAILS.split(','),
}




# Sensor IO keys — primary keys tried first, fallback to secondary if missing
IO_TEMP      = 'io10800'  # primary temp:     value / 100 = °C
IO_TEMP2     = 'io25'     # fallback temp:    value / 100 = °C
IO_HUMIDITY  = 'io10804'  # primary humidity: raw = %RH
IO_HUMIDITY2 = 'io86'     # fallback humidity: raw = %RH

# ── Colours / styles ───────────────────────────────────────────────────────────
FONT_NAME   = "Arial"
C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_SUBHDR_BG = "2E75B6"
C_SUBHDR_FG = "FFFFFF"
C_LABEL_BG  = "D6E4F0"
C_ALT_ROW   = "EBF3FA"
C_RED_BG    = "FFE0E0"
C_RED_FG    = "C00000"
C_DOOR_OPEN = "FFD7D7"
C_BORDER    = "8EA9C1"

thin = Side(style="thin", color=C_BORDER)

def _border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color, name=FONT_NAME)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def sanitize_filename(name):
    return _re.sub(r'[\\/*?:"<>|]', '_', name).strip() or 'unknown'


# ═══════════════════════════════════════════════════════════════════════════════
#  10-MINUTE RESAMPLING  (for bottom data table only)
# ═══════════════════════════════════════════════════════════════════════════════

def resample_10min(temp_data, humidity_data):
    """
    Bucket raw readings into 10-minute windows.
    Each bucket is labelled by its window start (floor to 10 min).
    Returns a list of dicts: {timestamp, temperature, humidity}
    temperature and humidity are averages of all readings in that window.
    """
    temp_buckets = defaultdict(list)
    hum_buckets  = defaultdict(list)

    for r in temp_data:
        ts = r['timestamp']
        # floor to nearest 10-minute mark
        bucket = ts.replace(minute=(ts.minute // 10) * 10, second=0, microsecond=0)
        temp_buckets[bucket].append(r['temperature'])

    for r in humidity_data:
        ts = r['timestamp']
        bucket = ts.replace(minute=(ts.minute // 10) * 10, second=0, microsecond=0)
        hum_buckets[bucket].append(r['humidity'])

    all_buckets = sorted(set(temp_buckets) | set(hum_buckets))

    result = []
    for bucket in all_buckets:
        temps = temp_buckets.get(bucket)
        hums  = hum_buckets.get(bucket)
        result.append({
            'timestamp':   bucket,
            'temperature': sum(temps) / len(temps) if temps else None,
            'humidity':    sum(hums)  / len(hums)  if hums  else None,
        })

    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  GPS API DATA FETCHING
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_vehicle_data(device_id, start_str, end_str, api_key):
    url = (
        f"https://fms2.gpsbox.mn/api/api.php"
        f"?api=user&key={api_key}"
        f"&cmd=OBJECT_GET_MESSAGES,{device_id},{start_str},{end_str},0.01"
    )
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    return response.json()


def parse_api_response(json_data, plate_number):
    """
    Parse raw GPS API messages into temperature, humidity, and door lists.
    Temp & humidity filtered to each day's active driving window:
    first non-zero speed -> last non-zero speed per day.
    """
    storage_temp = []
    humidity     = []
    door         = []
    last_valid_temp = None
    last_valid_hum  = None

    # Pass 1: build per-day active windows from speed (entry[5])
    daily_first = {}
    daily_last  = {}
    for entry in json_data:
        try:
            ts    = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
            speed = float(entry[5]) if len(entry) > 5 else 0.0
        except Exception:
            continue
        if speed > 0:
            d = ts.date()
            if d not in daily_first:
                daily_first[d] = ts
            daily_last[d] = ts

    # Pass 2: parse sensors, filter temp/hum to active window
    for entry in json_data:
        try:
            ts = datetime.strptime(entry[0], '%Y-%m-%d %H:%M:%S') + timedelta(hours=8)
            io = entry[6] if len(entry) > 6 and isinstance(entry[6], dict) else {}
        except Exception:
            continue

        d         = ts.date()
        in_window = (d in daily_first and daily_first[d] <= ts <= daily_last[d])

        # Temperature
        temp_key = IO_TEMP if IO_TEMP in io else (IO_TEMP2 if IO_TEMP2 in io else None)
        if temp_key:
            raw = float(io[temp_key])
            if raw == 250:
                temp = last_valid_temp
            else:
                temp = raw / 100.0
                last_valid_temp = temp
            if temp is not None and in_window:
                storage_temp.append({'timestamp': ts, 'temperature': temp})

        # Humidity
        hum_key = IO_HUMIDITY if IO_HUMIDITY in io else (IO_HUMIDITY2 if IO_HUMIDITY2 in io else None)
        if hum_key:
            raw = float(io[hum_key])
            hum = last_valid_hum if raw == 250 else raw
            if raw != 250:
                last_valid_hum = raw
            if hum is not None and in_window:
                humidity.append({'timestamp': ts, 'humidity': hum})

    print(f"  {plate_number}: temp={len(storage_temp)}pts  "
          f"hum={len(humidity)}pts  door={len(door)}pts  "
          f"active_days={len(daily_first)}")
    return storage_temp, humidity, door


# ═══════════════════════════════════════════════════════════════════════════════
#  STATISTICS
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_statistics(temp_data):
    if not temp_data:
        return None
    temps = [r['temperature'] for r in temp_data]
    dh, R = 83.144, 0.008314
    tk  = [t + 273.15 for t in temps]
    mkt = (dh / R) / (sum(dh / (R * t) for t in tk) / len(tk)) - 273.15
    return {
        'highest':      max(temps),
        'lowest':       min(temps),
        'average':      sum(temps) / len(temps),
        'mkt':          mkt,
        'data_points':  len(temps),
        'start_time':   temp_data[0]['timestamp'],
        'stop_time':    temp_data[-1]['timestamp'],
        'elapsed_time': temp_data[-1]['timestamp'] - temp_data[0]['timestamp'],
    }

def calculate_humidity_statistics(hum_data):
    if not hum_data:
        return None
    h = [r['humidity'] for r in hum_data]
    return {'highest': max(h), 'lowest': min(h),
            'average': sum(h) / len(h), 'data_points': len(h)}

def calculate_door_statistics(door_data):
    if not door_data:
        return None
    events, open_count, total_secs, open_at = [], 0, 0.0, None
    if door_data[0]['state'] == 0:
        open_at = door_data[0]['timestamp']
        open_count += 1
    for i in range(1, len(door_data)):
        prev = door_data[i-1]['state']
        curr = door_data[i]['state']
        ts   = door_data[i]['timestamp']
        if prev == 1 and curr == 0:
            open_at = ts; open_count += 1
        elif prev == 0 and curr == 1 and open_at:
            dur = (ts - open_at).total_seconds()
            total_secs += dur
            events.append({'opened_at': open_at, 'closed_at': ts, 'duration_secs': dur})
            open_at = None
    if open_at:
        dur = (door_data[-1]['timestamp'] - open_at).total_seconds()
        total_secs += dur
        events.append({'opened_at': open_at, 'closed_at': None, 'duration_secs': dur})
    return {'open_count': open_count, 'total_open_secs': total_secs,
            'events': events, 'data_points': len(door_data)}

def fmt_duration(secs):
    secs = int(secs)
    h, rem = divmod(secs, 3600); m, s = divmod(rem, 60)
    if h:   return f"{h}h {m}m {s}s"
    elif m: return f"{m}m {s}s"
    return f"{s}s"

def format_elapsed_time(td):
    if not td: return '0d 0h 0m'
    s = int(td.total_seconds())
    return f"{s//86400}d {(s%86400)//3600}h {(s%3600)//60}m"


# ═══════════════════════════════════════════════════════════════════════════════
#  CELL HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _write_cell(ws, row, col, value, bold=False, size=10, fg="000000",
                bg=None, halign="left", border=True, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=fg)
    cell.alignment = _align(h=halign, wrap=wrap)
    if bg:
        cell.fill = _fill(bg)
    if border:
        cell.border = _border()
    return cell

def _section_header(ws, row, col, text, span, bg=C_HEADER_BG, fg=C_HEADER_FG):
    _write_cell(ws, row, col, text, bold=True, size=11, fg=fg, bg=bg,
                halign="left", wrap=True)
    ws.row_dimensions[row].height = 20
    for c in range(col + 1, col + span):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)

def _spacer(ws, row, height=10):
    ws.row_dimensions[row].height = height


# ═══════════════════════════════════════════════════════════════════════════════
#  PLATE SHEET BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_plate_sheet(wb, plate_number, analysis):
    safe_name = sanitize_filename(plate_number)[:31]
    ws = wb.create_sheet(title=safe_name)
    ws.sheet_view.showGridLines = False

    col_widths = [16, 8, 8, 2, 16, 8, 8, 2, 16, 8, 8, 2]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    temp_data     = analysis['storage_temp_data']
    humidity_data = analysis.get('humidity_data', [])
    door_data     = analysis.get('door_data', [])

    stats      = calculate_statistics(temp_data)
    hum_stats  = calculate_humidity_statistics(humidity_data)
    door_stats = calculate_door_statistics(door_data)

    if not stats:
        ws['A1'] = f"Тухайн өдөр ямар ч хөдөлгөөн хийгээгүй байна {plate_number}"
        return

    # ── Build 10-minute resampled data for the bottom table ───────────────────
    table_data = resample_10min(temp_data, humidity_data)

    # ── Chart data on hidden sheet ─────────────────────────────────────────────
    ds_name = f"_d_{safe_name}"[:31]
    ds = wb.create_sheet(title=ds_name)
    ds.sheet_state = 'hidden'

    humidity_dict = {r['timestamp']: r['humidity'] for r in humidity_data}

    step    = max(1, len(temp_data) // 500)
    sampled = temp_data[::step]
    n_pts   = len(sampled)

    ds.cell(row=1, column=1, value="Цаг")
    ds.cell(row=1, column=2, value="Temp °C")
    ds.cell(row=1, column=3, value="Humidity %RH")
    for i, rd in enumerate(sampled, 2):
        is_first = (i == 2)
        is_last  = (i == n_pts + 1)
        label = rd['timestamp'].strftime('%Y/%m/%d %H:%M') if (is_first or is_last) else ''
        ds.cell(row=i, column=1, value=label)
        ds.cell(row=i, column=2, value=round(rd['temperature'], 1))
        hv = humidity_dict.get(rd['timestamp'])
        ds.cell(row=i, column=3, value=round(hv, 1) if hv is not None else None)

    # Build chart
    chart = LineChart()
    chart.title        = f"{plate_number} — Температур & Чийгшил"
    chart.y_axis.title = "Temperature (°C)"
    chart.y_axis.axId  = 100
    chart.x_axis.axId  = 10
    chart.x_axis.tickLblPos = "low"
    chart.style        = 10
    chart.width        = 22
    chart.height       = 10

    cats  = Reference(ds, min_col=1, min_row=2, max_row=n_pts + 1)
    t_ref = Reference(ds, min_col=2, min_row=1, max_row=n_pts + 1)
    chart.add_data(t_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "1F4E79"
    chart.series[0].graphicalProperties.line.width     = 20000

    if hum_stats:
        chart2 = LineChart()
        chart2.y_axis.title   = "Humidity (RH%)"
        chart2.y_axis.axId    = 200
        chart2.y_axis.crosses = "max"
        chart2.y_axis.scaling.min = 0
        _hmax = hum_stats['highest']
        chart2.y_axis.scaling.max = (int(_hmax // 10) + 1) * 10
        chart2.x_axis.axId    = 10
        chart2.x_axis.tickLblPos = "low"
        h_ref = Reference(ds, min_col=3, min_row=1, max_row=n_pts + 1)
        chart2.add_data(h_ref, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.series[0].graphicalProperties.line.solidFill = "00BFFF"
        chart2.series[0].graphicalProperties.line.width     = 15000
        chart += chart2

    # ── Title rows ─────────────────────────────────────────────────────────────
    row = 1
    ws.row_dimensions[row].height = 26
    _write_cell(ws, row, 1, "Дата тайлан — " + plate_number,
                bold=True, size=16, fg=C_HEADER_BG, border=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1
    _write_cell(ws, row, 1,
                f"Машины дугаар: {plate_number}   |   "
                f"{stats['start_time'].strftime('%Y/%m/%d')} – "
                f"{stats['stop_time'].strftime('%Y/%m/%d')}",
                bold=False, size=10, fg="555555", border=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1

    # Place chart, reserve 20 rows
    ws.add_chart(chart, f"A{row}")
    CHART_ROWS = 20
    row += CHART_ROWS

    for _ in range(3):
        _spacer(ws, row); row += 1

    # ── Device info panels ─────────────────────────────────────────────────────
    _section_header(ws, row, 1, "📋  Мэдээлэл (Device Info)", 12)
    row += 1

    elapsed  = format_elapsed_time(stats['elapsed_time'])
    cfg_rows = [
        ("Товч мэдээлэл",  ""),
        ("Эхэлсэн цаг",    stats['start_time'].strftime('%Y/%m/%d %H:%M:%S')),
        ("Эцсийн цаг",     stats['stop_time'].strftime('%Y/%m/%d %H:%M:%S')),
        ("Нийт цаг",       elapsed),
        ("Time Zone",      "+08:00"),
        ("Start Delay",    "0 min"),
    ]
    tmp_rows = [
        ("Температур мэдээлэл", ""),
        ("Хамгийн өндөр",  f"{stats['highest']:.2f} °C"),
        ("Хамгийн бага",   f"{stats['lowest']:.2f} °C"),
        ("Дундаж",         f"{stats['average']:.2f} °C"),
        ("MKT",            f"{stats['mkt']:.2f} °C"),
        ("Датаны тоо",     stats['data_points']),
    ]
    hum_info = [
        ("Чийгшил мэдээлэл", ""),
        ("Хамгийн өндөр",  f"{hum_stats['highest']:.2f} %" if hum_stats else "-"),
        ("Хамгийн бага",   f"{hum_stats['lowest']:.2f} %"  if hum_stats else "-"),
        ("Дундаж",         f"{hum_stats['average']:.2f} %"  if hum_stats else "-"),
        ("Alarm At",       "-"),
        ("Датаны тоо",     hum_stats['data_points'] if hum_stats else "-"),
    ]

    for i, ((l1, v1), (l2, v2), (l3, v3)) in enumerate(
            zip(cfg_rows, tmp_rows, hum_info)):
        r   = row + i
        hdr = (i == 0)
        bg  = C_SUBHDR_BG if hdr else C_LABEL_BG
        fg  = C_SUBHDR_FG if hdr else "000000"
        _write_cell(ws, r, 1,  l1, bold=hdr, fg=fg, bg=bg)
        _write_cell(ws, r, 2,  v1)
        _write_cell(ws, r, 5,  l2, bold=hdr, fg=fg, bg=bg)
        _write_cell(ws, r, 6,  v2)
        _write_cell(ws, r, 9,  l3, bold=hdr, fg=fg, bg=bg)
        _write_cell(ws, r, 10, v3)
        if hdr:
            ws.merge_cells(start_row=r, start_column=1,  end_row=r, end_column=3)
            ws.merge_cells(start_row=r, start_column=5,  end_row=r, end_column=6)
            ws.merge_cells(start_row=r, start_column=9,  end_row=r, end_column=10)

    row += len(cfg_rows)
    _spacer(ws, row); row += 1

    # ── Door stats ─────────────────────────────────────────────────────────────
    if door_stats:
        _section_header(ws, row, 1, "🚪  Хаалга мэдээлэл (Door Info)", 12)
        row += 1
        for label, value in [
            ("Нийт нээсэн тоо",         f"{door_stats['open_count']} удаа"),
            ("Нийт нээлттэй байсан цаг", fmt_duration(door_stats['total_open_secs'])),
            ("Датаны тоо",               str(door_stats['data_points'])),
        ]:
            _write_cell(ws, row, 1, label, bold=True, bg=C_LABEL_BG)
            for c in range(2, 4):
                cell = ws.cell(row=row, column=c)
                cell.fill = _fill(C_LABEL_BG); cell.border = _border()
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            _write_cell(ws, row, 4, value)
            for c in range(5, 7):
                cell = ws.cell(row=row, column=c)
                cell.border = _border()
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
            row += 1

        if door_stats['events']:
            _spacer(ws, row); row += 1
            DOOR_COLS = [
                (1,  1,  "#"),
                (2,  4,  "Нээгдсэн цаг"),
                (5,  8,  "Хаагдсан цаг"),
                (9,  12, "Нээлттэй байсан"),
            ]
            for start, end, h in DOOR_COLS:
                _write_cell(ws, row, start, h, bold=True,
                            bg=C_SUBHDR_BG, fg=C_SUBHDR_FG, halign="center")
                if end > start:
                    for c in range(start+1, end+1):
                        cell = ws.cell(row=row, column=c)
                        cell.fill = _fill(C_SUBHDR_BG); cell.border = _border()
                    ws.merge_cells(start_row=row, start_column=start,
                                   end_row=row, end_column=end)
            row += 1
            for idx, ev in enumerate(door_stats['events'], 1):
                opened = ev['opened_at'].strftime('%Y/%m/%d %H:%M:%S')
                closed = (ev['closed_at'].strftime('%Y/%m/%d %H:%M:%S')
                          if ev['closed_at'] else "—")
                dur    = (fmt_duration(ev['duration_secs'])
                          if ev['duration_secs'] is not None else "—")
                bg_ev  = (C_DOOR_OPEN if not ev['closed_at']
                          else (C_ALT_ROW if idx % 2 == 0 else "FFFFFF"))
                for start, end, val in [
                    (1,  1,  str(idx)),
                    (2,  4,  opened),
                    (5,  8,  closed),
                    (9,  12, dur),
                ]:
                    _write_cell(ws, row, start, val, bg=bg_ev, halign="center")
                    if end > start:
                        for c in range(start+1, end+1):
                            cell = ws.cell(row=row, column=c)
                            cell.fill = _fill(bg_ev); cell.border = _border()
                        ws.merge_cells(start_row=row, start_column=start,
                                       end_row=row, end_column=end)
                row += 1

        _spacer(ws, row); row += 1

    # ── Temperature & Humidity data table (10-min averages) ───────────────────
    _section_header(ws, row, 1,
                    "🌡️  Температур & Чийгшил өгөгдөл — 10 мин дундаж "
                    "(Temperature & Humidity Data — 10 min avg)",
                    12)
    row += 1

    COL_SETS   = [(1, 2, 3), (5, 6, 7), (9, 10, 11)]
    HDR_LABELS = ["Цаг", "°C", "%RH"]

    for cs in COL_SETS:
        for ci, h in zip(cs, HDR_LABELS):
            _write_cell(ws, row, ci, h, bold=True,
                        bg=C_SUBHDR_BG, fg=C_SUBHDR_FG, halign="center")
    row += 1

    total       = len(table_data)
    rows_needed = max(1, (total + 2) // 3)

    for i in range(rows_needed):
        alt = (i % 2 == 1)
        for set_idx, cs in enumerate(COL_SETS):
            data_idx = set_idx * rows_needed + i
            empty_bg = C_ALT_ROW if alt else "FFFFFF"
            if data_idx >= total:
                for ci in cs:
                    c = ws.cell(row=row, column=ci, value="")
                    c.border = _border(); c.fill = _fill(empty_bg)
                continue

            reading = table_data[data_idx]
            t  = reading['temperature']
            hv = reading['humidity']

            # out-of-range check (only if temp exists)
            is_oor = (t is not None) and (t <= 15 or t >= 25)
            row_bg = C_RED_BG if is_oor else (C_ALT_ROW if alt else "FFFFFF")
            t_fg   = C_RED_FG if is_oor else "000000"

            # Timestamp cell — shows bucket start time (HH:MM, no seconds)
            ts_c = ws.cell(row=row, column=cs[0],
                           value=reading['timestamp'].strftime('%m/%d %H:%M'))
            ts_c.font = _font(size=8); ts_c.alignment = _align(h="center")
            ts_c.border = _border(); ts_c.fill = _fill(row_bg)

            # Temperature cell
            t_val = round(t, 1) if t is not None else "-"
            t_c = ws.cell(row=row, column=cs[1], value=t_val)
            t_c.font = _font(size=8, color=t_fg, bold=is_oor)
            t_c.alignment = _align(h="center"); t_c.border = _border()
            t_c.fill = _fill(C_RED_BG if is_oor else row_bg)
            if isinstance(t_val, float):
                t_c.number_format = '0.0'

            # Humidity cell
            h_val = round(hv, 1) if hv is not None else "-"
            h_c   = ws.cell(row=row, column=cs[2], value=h_val)
            h_c.font = _font(size=8); h_c.alignment = _align(h="center")
            h_c.border = _border(); h_c.fill = _fill(row_bg)
            if isinstance(h_val, float):
                h_c.number_format = '0.0'

        row += 1

    _spacer(ws, row); row += 1

    # ── Out-of-range table (still uses raw data for full accuracy) ─────────────
    oor = [r for r in temp_data if r['temperature'] <= 15 or r['temperature'] >= 25]
    if oor:
        _section_header(ws, row, 1,
                        "⚠️  Хязгаараас гарсан температур (Out-of-Range)", 12,
                        bg="C00000", fg="FFFFFF")
        row += 1
        for cs in COL_SETS:
            for ci, h in zip(cs, HDR_LABELS):
                _write_cell(ws, row, ci, h, bold=True,
                            bg="FF4444", fg="FFFFFF", halign="center")
        row += 1

        humidity_dict_raw = {r['timestamp']: r['humidity'] for r in humidity_data}
        oor_rows = max(1, (len(oor) + 2) // 3)
        for i in range(oor_rows):
            alt = (i % 2 == 1)
            bg  = "FFD0D0" if alt else C_RED_BG
            for set_idx, cs in enumerate(COL_SETS):
                data_idx = set_idx * oor_rows + i
                if data_idx >= len(oor):
                    for ci in cs:
                        c = ws.cell(row=row, column=ci, value="")
                        c.border = _border(); c.fill = _fill(bg)
                    continue
                reading = oor[data_idx]
                t  = reading['temperature']
                hv = humidity_dict_raw.get(reading['timestamp'])

                ts_c = ws.cell(row=row, column=cs[0],
                               value=reading['timestamp'].strftime('%m/%d %H:%M:%S'))
                ts_c.font = _font(size=8); ts_c.alignment = _align(h="center")
                ts_c.border = _border(); ts_c.fill = _fill(bg)

                t_c = ws.cell(row=row, column=cs[1], value=round(t, 1))
                t_c.font = _font(size=8, color=C_RED_FG, bold=True)
                t_c.alignment = _align(h="center"); t_c.border = _border()
                t_c.fill = _fill(bg); t_c.number_format = '0.0'

                h_val = round(hv, 1) if hv is not None else "-"
                h_c   = ws.cell(row=row, column=cs[2], value=h_val)
                h_c.font = _font(size=8); h_c.alignment = _align(h="center")
                h_c.border = _border(); h_c.fill = _fill(bg)
                if isinstance(h_val, float):
                    h_c.number_format = '0.0'
            row += 1

    ws.freeze_panes = "A3"
    print(f"  Sheet built: {safe_name}")


# ═══════════════════════════════════════════════════════════════════════════════
#  SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_summary_sheet(wb, plate_analyses):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    widths = [26, 22, 18, 18, 18, 18, 18, 18, 20, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    ws.row_dimensions[row].height = 26
    _write_cell(ws, row, 1, "Дата тайлан — Нэгтгэл (Summary Report)",
                bold=True, size=16, fg=C_HEADER_BG, border=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    row += 2

    headers = ["Машины дугаар", "Эхэлсэн цаг", "Хамгийн өндөр °C",
               "Хамгийн бага °C", "Дундаж °C", "MKT °C",
               "Хамгийн өндөр %RH", "Дундаж %RH",
               "Хаалга нээсэн тоо", "Нийт нээлттэй", "Датаны тоо"]
    for c, h in enumerate(headers, 1):
        _write_cell(ws, row, c, h, bold=True,
                    bg=C_HEADER_BG, fg=C_HEADER_FG, halign="center")
    row += 1

    for i, (plate, analysis) in enumerate(plate_analyses.items()):
        stats      = calculate_statistics(analysis['storage_temp_data'])
        hum_stats  = calculate_humidity_statistics(analysis.get('humidity_data', []))
        door_stats = calculate_door_statistics(analysis.get('door_data', []))
        if not stats:
            continue
        bg = C_ALT_ROW if i % 2 else "FFFFFF"
        values = [
            plate,
            stats['start_time'].strftime('%Y/%m/%d %H:%M'),
            round(stats['highest'], 2),
            round(stats['lowest'],  2),
            round(stats['average'], 2),
            round(stats['mkt'],     2),
            round(hum_stats['highest'], 2) if hum_stats else "-",
            round(hum_stats['average'], 2) if hum_stats else "-",
            door_stats['open_count']                    if door_stats else "-",
            fmt_duration(door_stats['total_open_secs']) if door_stats else "-",
            stats['data_points'],
        ]
        for c, v in enumerate(values, 1):
            _write_cell(ws, row, c, v, bg=bg, halign="center")
        row += 1

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

def send_email_with_attachment(sender_email, sender_password, receiver_emails,
                               subject, message, attachment_path):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.application import MIMEApplication
    from email.mime.text import MIMEText

    if isinstance(receiver_emails, str):
        receiver_emails = [receiver_emails]
    receiver_emails = [e.strip() for e in receiver_emails if e.strip()]
    if not receiver_emails:
        print("No valid receiver emails provided"); return False

    try:
        msg = MIMEMultipart()
        msg['From']    = sender_email
        msg['To']      = ', '.join(receiver_emails)
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))
        with open(attachment_path, 'rb') as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
            part['Content-Disposition'] = (
                f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(part)
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(msg)
        print(f"Email sent to: {', '.join(receiver_emails)}")
        return True
    except Exception as e:
        print(f"Error sending email: {e}"); return False


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    today     = datetime.now()
    Yesterday = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    Tooday    = today.replace(hour=0, minute=0, second=0, microsecond=0)

    start_date = Yesterday - timedelta(hours=8)
    end_date   = Tooday    - timedelta(hours=8)

    start_str = start_date.strftime('%Y-%m-%d %H:%M')
    end_str   = end_date.strftime('%Y-%m-%d %H:%M')

    print(f"Fetching data: {start_str}  →  {end_str}\n")

    plate_analyses = {}

    for device_id, plate_number in CONFIG['VEHICLES'].items():
        print(f"Fetching {plate_number} ({device_id})…")
        try:
            raw = fetch_vehicle_data(device_id, start_str, end_str, CONFIG['GPS_API_KEY'])
            temp_data, hum_data, door_data = parse_api_response(raw, plate_number)
            plate_analyses[plate_number] = {
                'storage_temp_data': temp_data,
                'humidity_data':     hum_data,
                'door_data':         door_data,
            }
        except Exception as e:
            print(f"  ERROR for {plate_number}: {e}")

    if not plate_analyses:
        print("No data fetched for any vehicle."); return

    os.makedirs('reports', exist_ok=True)
    output_file = 'reports/sensor_report.xlsx'

    wb = Workbook()
    build_summary_sheet(wb, plate_analyses)
    for plate_number, analysis in plate_analyses.items():
        build_plate_sheet(wb, plate_number, analysis)
    wb.save(output_file)
    print(f"\nReport saved: {output_file}")

    # Send email
    send_email_with_attachment(
        CONFIG['SENDER_EMAIL'],
        CONFIG['SENDER_PASSWORD'],
        CONFIG['RECEIVER_EMAILS'],
        f"Sensor Report — {Tooday}",
        f"Attached is the temperature & humidity report for all vehicles.\n"
        f"Period: {Yesterday} → {Tooday}",
        output_file,
    )


if __name__ == "__main__":
    main()
